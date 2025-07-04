from flask import Blueprint, render_template, request, redirect, url_for
from flask import session
from . import db
from .models import Product, Transaction
from openpyxl import load_workbook

bp = Blueprint('pos', __name__)


def get_cart():
    return session.setdefault('cart', {})


def cart_details():
    cart = get_cart()
    items = []
    total = 0
    for pid, data in cart.items():
        product = Product.query.get(int(pid))
        if not product:
            continue
        if isinstance(data, int):
            qty = data
            price = product.price
        else:
            qty = int(data.get('qty', 1))
            price = float(data.get('price', product.price))
        subtotal = price * qty
        items.append({'product': product, 'qty': qty, 'price': price, 'subtotal': subtotal})
        total += subtotal
    return items, total


@bp.app_context_processor
def inject_cart():
    items, total = cart_details()
    return {'cart_items': items, 'cart_total': total}


@bp.route('/')
def index():
    products = Product.query.all()
    return render_template('products.html', products=products)


@bp.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        file = request.files.get('file')
        if file:
            wb = load_workbook(file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                name, price = row[:2]
                if name and price is not None:
                    product = Product(name=str(name), price=float(price))
                    db.session.add(product)
            db.session.commit()
            return redirect(url_for('pos.index'))
    return render_template('upload.html')


@bp.route('/add/<int:product_id>')
def add_to_cart(product_id):
    cart = get_cart()
    key = str(product_id)
    product = Product.query.get_or_404(product_id)
    if key in cart:
        item = cart[key]
        if isinstance(item, dict):
            item['qty'] += 1
        else:
            cart[key] = {'qty': item + 1, 'price': product.price}
    else:
        cart[key] = {'qty': 1, 'price': product.price}
    session.modified = True
    return redirect(url_for('pos.index'))


@bp.route('/cart')
def view_cart():
    items, total = cart_details()
    return render_template('cart.html', items=items, total=total)


@bp.route('/update/<int:product_id>', methods=['POST'])
def update_cart(product_id):
    cart = get_cart()
    key = str(product_id)
    qty = int(request.form.get('qty', 1))
    price = float(request.form.get('price', 0))
    cart[key] = {'qty': qty, 'price': price}
    session.modified = True
    next_url = request.form.get('next') or url_for('pos.view_cart')
    return redirect(next_url)


@bp.route('/checkout', methods=['POST'])
def checkout():
    cart = get_cart()
    items = []
    total = 0
    for pid, data in cart.items():
        product = Product.query.get(int(pid))
        if product:
            if isinstance(data, int):
                qty = data
                price = product.price
            else:
                qty = int(data.get('qty', 1))
                price = float(data.get('price', product.price))
            items.append({
                'id': product.id,
                'name': product.name,
                'price': price,
                'qty': qty
            })
            total += price * qty
    if items:
        tx = Transaction(total=total)
        tx.set_items(items)
        db.session.add(tx)
        db.session.commit()
    session['cart'] = {}
    return redirect(url_for('pos.index'))
