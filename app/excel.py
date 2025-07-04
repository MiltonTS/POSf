from openpyxl import load_workbook
from .models import Product
from . import db


def load_products_from_excel(file_path, name_column=0, price_column=1):
    """Load products from an Excel file.

    Parameters
    ----------
    file_path : str
        Path to the Excel file.
    name_column : int, default 0
        Index of the column with product names.
    price_column : int, default 1
        Index of the column with prices.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        if len(row) <= max(name_column, price_column):
            continue
        name = row[name_column]
        price = row[price_column]
        if name and price is not None:
            product = Product(name=str(name), price=float(price))
            db.session.add(product)
    db.session.commit()
